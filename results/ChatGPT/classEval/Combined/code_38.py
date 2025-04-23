import openpyxl


class ExcelProcessor:
    """
    This class processes Excel files, including reading and writing data,
    as well as performing specific operations and saving as a new Excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reads data from an Excel file.
        
        :param file_name: str, Excel file name to read
        :return: list of tuples, Data in Excel or None if file_name is empty
        """
        if not file_name:
            return None
        
        workbook = openpyxl.load_workbook(file_name, data_only=True)
        sheet = workbook.active
        data = [tuple(cell.value for cell in row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return data

    def write_excel(self, data, file_name):
        """
        Writes data to the specified Excel file.
        
        :param data: list of tuples, Data to be written
        :param file_name: str, Excel file name to write to
        :return: int, 1 for success, 0 for failure
        """
        if not file_name or not data:
            return 0
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(file_name)
        workbook.close()
        return 1

    def process_excel_data(self, column_index, source_file_name):
        """
        Changes the specified column in the Excel file to uppercase.
        
        :param column_index: int, The serial number of the column to change
        :param source_file_name: str, source file name
        :return: (int, str), The first is the return value of write_excel,
                 while the second is the saved file name of the processed data
        """
        data = self.read_excel(source_file_name)
        if data is None or column_index < 0 or column_index >= len(data[0]):
            return 0, ""

        processed_data = []
        header = data[0] + (data[0][column_index].upper(),)
        processed_data.append(header)

        for row in data[1:]:
            new_row = list(row) + (row[column_index],)
            processed_data.append(new_row)

        output_file_name = f"processed_{source_file_name}"
        success = self.write_excel(processed_data, output_file_name)
        return success, output_file_name