import openpyxl


class ExcelProcessor:
    """
    A class for processing Excel files, including reading and writing data, and performing specific operations.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Read data from an Excel file.
        :param file_name: str, Excel file name to read.
        :return: list of tuples, Data in Excel.
        """
        if not file_name:
            print("File name is empty.")
            return None

        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            return data
        except Exception as e:
            print(f"Error reading {file_name}: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file.
        :param data: list of tuples, Data to be written.
        :param file_name: str, Excel file name to write to.
        :return: int, 1 for success, 0 for failure.
        """
        if not file_name or not data:
            print("File name or data to write is empty.")
            return 0

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing to {file_name}: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase.
        :param N: int, The serial number of the column to change.
        :param save_file_name: str, source file name.
        :return: (int, str), Status of write operation and the saved file name.
        """
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]):
            return 0, ""

        new_data = []
        for row in data:
            new_row = list(row) + [row[N].upper() if isinstance(row[N], str) else row[N]]
            new_data.append(tuple(new_row))

        output_file_name = f"processed_{save_file_name}"
        success = self.write_excel(new_data, output_file_name)
        return success, output_file_name