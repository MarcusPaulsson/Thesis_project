import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        if not file_name:
            return None
        data = []
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        if not file_name:
            return 0
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(file_name)
        return 1

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        if not data or N < 0 or N >= len(data[0]):
            return 0, ""
        
        new_data = []
        header = list(data[0]) + [data[0][N].upper()]
        new_data.append(tuple(header))
        
        for row in data[1:]:
            new_row = list(row) + [row[N]]
            new_data.append(tuple(new_row))

        output_file_name = f'processed_{save_file_name}'
        success = self.write_excel(new_data, output_file_name)
        return success, output_file_name