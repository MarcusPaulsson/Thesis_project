import openpyxl


class ExcelProcessor:
    """
    A class for processing Excel files, including reading and writing Excel data,
    as well as processing specific operations and saving as a new Excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Read data from an Excel file.
        
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file.
        
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: int, 1 for successful writing, 0 for failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            for row in data:
                sheet.append(row)

            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing Excel file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase.
        
        :param N: int, The serial number of the column to change (1-indexed)
        :param save_file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel,
                 while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)

        # Convert the specified column to uppercase
        for i in range(1, len(data)):
            if len(data[i]) >= N:
                data[i] = list(data[i])  # Convert tuple to list to modify
                data[i][N - 1] = str(data[i][N - 1]).upper()  # Change to uppercase
                data[i] = tuple(data[i])  # Convert back to tuple

        output_file_name = f"processed_{save_file_name}"
        success = self.write_excel(data, output_file_name)

        return success, output_file_name