import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1  # Success
        except Exception:
            return 0  # Failure

    def process_excel_data(self, N, file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change (1-indexed)
        :param file_name: str, source file name
        :return: (int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(file_name)
        processed_data = []

        for row in data:
            if N <= len(row):
                # Change the specified column to uppercase
                new_row = list(row)
                new_row[N - 1] = str(new_row[N - 1]).upper()
                processed_data.append(new_row)
            else:
                processed_data.append(row)

        output_file_name = f"processed_{file_name}"
        success = self.write_excel(processed_data, output_file_name)
        return success, output_file_name