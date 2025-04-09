import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name: str, Excel file name to read
        :return: list of data, Data in Excel
        """
        if not file_name:
            return None
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(tuple(row))
            return data
        except FileNotFoundError:
            return None
        except Exception:
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        if not file_name:
            return 0
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        if data is None:
            return 0, ""

        if N < 0 or N >= len(data[0]):
            return 0, ""

        processed_data = []
        header = list(data[0])
        header[N] = header[N].upper()
        processed_data.append(tuple(header))

        for row in data[1:]:
            new_row = list(row)
            if isinstance(new_row[N], str):
                new_row[N] = str(new_row[N]).upper()
            else:
                new_row[N] = str(new_row[N])
            processed_data.append(tuple(new_row))

        output_file_name = "processed_" + save_file_name
        success = self.write_excel(processed_data, output_file_name)
        return success, output_file_name