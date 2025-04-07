import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows():
                row_data = tuple(cell.value for cell in row)
                data.append(row_data)
            return data
        except FileNotFoundError:
            return None
        except openpyxl.utils.exceptions.InvalidFileException:
            return None
        except Exception as e:
            print(f"Error reading excel file: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        if not file_name:
            return 0
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row_data in data:
                sheet.append(row_data)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing to excel file: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        if N < 0:
            return 0
        try:
            data = self.read_excel(save_file_name)
            if data is None:
                return 0
            new_data = []
            for row in data:
                new_row = list(row)
                if N < len(new_row):
                    new_row.append(str(new_row[N]).upper())
                else:
                    return 0
                new_data.append(tuple(new_row))

            output_file_name = "processed_" + save_file_name
            success = self.write_excel(new_data, output_file_name)
            return success, output_file_name
        except FileNotFoundError:
            return 0
        except Exception as e:
            print(f"Error processing excel data: {e}")
            return 0