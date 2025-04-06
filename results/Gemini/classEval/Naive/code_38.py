import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return data
        except FileNotFoundError:
            print(f"Error: File '{file_name}' not found.")
            return None
        except Exception as e:
            print(f"An error occurred while reading the Excel file: {e}")
            return None

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            for row_data in data:
                sheet.append(row_data)

            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"An error occurred while writing to the Excel file: {e}")
            return 0

    def process_excel_data(self, N, file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        try:
            data = self.read_excel(file_name)
            if data is None:
                return 0, None

            processed_data = []
            for row in data:
                new_row = list(row)  # Convert tuple to list to allow modification
                if len(new_row) > N - 1:
                    if new_row[N - 1] is not None and isinstance(new_row[N - 1], str):
                        new_row[N - 1] = new_row[N - 1].upper()
                processed_data.append(tuple(new_row))  # Convert back to tuple

            save_file_name = f"processed_{file_name}"
            success = self.write_excel(processed_data, save_file_name)
            return success, save_file_name
        except Exception as e:
            print(f"An error occurred during processing: {e}")
            return 0, None

if __name__ == '__main__':
    processor = ExcelProcessor()

    # Example usage of write_excel
    new_data = [
        ('Name', 'Age', 'Country'),
        ('John', 25, 'USA'),
        ('Alice', 30, 'Canada'),
        ('Bob', 35, 'Australia'),
        ('Julia', 28, 'Germany')
    ]
    write_success = processor.write_excel(new_data, 'test_data.xlsx')
    print(f"Write Excel successful: {write_success}")

    # Example usage of process_excel_data
    success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
    print(f"Process Excel successful: {success}, Output file: {output_file}")